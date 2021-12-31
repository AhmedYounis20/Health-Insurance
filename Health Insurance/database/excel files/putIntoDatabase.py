import openpyxl
import mysql.connector


mydatabase=mysql.connector.connect(
    host='localhost',
    user='team',
    password='team',
    database='InsuranceCompany',
)
hospitals=openpyxl.load_workbook('hospitals.xlsx')
mycursor=mydatabase.cursor()
sql='insert into hospital(name,website,Country,region,city,street)values(%s,%s,%s,%s,%s,%s);'
values=[]
hos=hospitals.active

for i in range(2,hos.max_row+1):
    current=[]
    for j in range (1,hos.max_column+1):
        current+=[hos.cell(i,j).value]
    values.append(tuple(current))
    print(current)
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'hospitals inserted')


plans=openpyxl.load_workbook('plans.xlsx')
sql='insert into plan(type,description)values(%s,%s);'
values=[]
plans=plans.active

for i in range(2,plans.max_row+1):
    current=[]
    for j in range (1,plans.max_column+1):
        current+=[plans.cell(i,j).value]
    values.append(tuple(current))
    print(current)
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'plans inserted')


customers=openpyxl.load_workbook('customers.xlsx')
sql='insert into customer(holderid,planid,firstname,lastname,email,age,RegistrationDate,stuff)values(%s,%s,%s,%s,%s,%s,%s,%s);'
values=[]
customers=customers.active

for i in range(2,customers.max_row+1):
    current=[]
    for j in range (2,customers.max_column+1):
        current+=[customers.cell(i,j).value]
    print(current)
    values.append(tuple(current))
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'customers inserted')




hospitalplan=openpyxl.load_workbook('hospitalplan.xlsx')

sql='insert into hospitalplan(hospitalid,planid)values(%s,%s);'
values=[]
hospitalplan=hospitalplan.active

for i in range(2,hospitalplan.max_row+1):
    current=[]
    for j in range (1,hospitalplan.max_column+1):
        current+=[hospitalplan.cell(i,j).value]
    print(current)
    values.append(tuple(current))
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'hospital to plan relationships inserted')



contacts=openpyxl.load_workbook('contacts.xlsx')

sql='insert into hospitalcontacts(hospitalid,phone)values(%s,%s);'
values=[]
hospitalcontacts=contacts.active

for i in range(2,hospitalcontacts.max_row+1):
    current=[] 
    if hospitalcontacts.cell(i,1).value=="hospital":
        for j in range (2,hospitalcontacts.max_column+1):
            current+=[hospitalcontacts.cell(i,j).value]
        print(current)
        values.append(tuple(current))
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'hospital contacts inserted')



sql='insert into customercontact(customerid,phone)values(%s,%s);'
values=[]

for i in range(2,hospitalcontacts.max_row+1):
    current=[] 
    if hospitalcontacts.cell(i,1).value=="customer":
        for j in range (2,hospitalcontacts.max_column+1):
            current+=[hospitalcontacts.cell(i,j).value]
        print(current)
        values.append(tuple(current))
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'customer contacts inserted')




claims=openpyxl.load_workbook('claims.xlsx')

sql='insert into claim(hospitalid,customerid,approved,submittingdate,expense,description)values(%s,%s,%s,%s,%s,%s);'
values=[]
claims=claims.active

for i in range(2,claims.max_row+1):
    current=[] 
    for j in range (1,claims.max_column+1):
        current+=[claims.cell(i,j).value]
    print(current)
    values.append(tuple(current))
mycursor.executemany(sql,values);
print(mycursor.rowcount)
mydatabase.commit();
print(mycursor.rowcount,'claims inserted')


